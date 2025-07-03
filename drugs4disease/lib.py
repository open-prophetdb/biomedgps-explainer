import os
import logging
import numpy as np
import pandas as pd
import torch as th
from gprofiler import GProfiler
import networkx as nx
import requests
import json
from openpyxl import load_workbook
from typing import List, Tuple, Union, Dict

# conda install graph-tool -c conda-forge
import graph_tool.all as gt
import torch.nn.functional as fn


# function: calculate the cosine similarity between two vectors
def cosine_similarity(vec1: list, vec2: list) -> float:
    """
    Calculate the cosine similarity between two vectors.

    Args:
        vec1 (list): vector 1
        vec2 (list): vector 2

    Returns:
        float: cosine similarity
    """
    return np.dot(vec1, vec2) / (np.linalg.norm(vec1) * np.linalg.norm(vec2))


def euclidean_distance(vec1: list, vec2: list) -> float:
    """
    Calculate the Euclidean distance between two vectors.

    Args:
        vec1 (list): vector 1
        vec2 (list): vector 2

    Returns:
        float: Euclidean distance
    """
    return np.linalg.norm(vec1 - vec2)


def manhattan_distance(vec1: list, vec2: list) -> float:
    """
    Calculate the Manhattan distance between two vectors.

    Args:
        vec1 (list): vector 1
        vec2 (list): vector 2

    Returns:
        float: Manhattan distance
    """
    return np.linalg.norm(vec1 - vec2, ord=1)


def median(lst: list) -> float:
    """Calculate the median of the list."""
    return np.median(np.array(lst))


def mean(lst: list) -> float:
    """Calculate the mean of the list."""
    if np.isnan(lst).any():
        return 0

    return np.mean(np.array(lst))


def get_logger(name: str, log_file: str = None) -> logging.Logger:
    """Get logger.

    Args:
        name (str): logger name
        log_file (str, optional): log file. Defaults to None.

    Returns:
        logger: logger
    """
    logger = logging.getLogger(name)
    logger.setLevel(logging.DEBUG)
    formatter = logging.Formatter(
        "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
    )

    if log_file:
        fh = logging.FileHandler(log_file)
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(formatter)
        logger.addHandler(fh)

    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)
    ch.setFormatter(formatter)
    logger.addHandler(ch)

    return logger


logger = get_logger("lib")


def get_genes(df):
    genes = df["source_id"].tolist() + df["target_id"].tolist()
    filtered_genes = filter(lambda x: x.startswith("ENTREZ:"), genes)
    return list(set(filtered_genes))


def get_gene_names(df):
    source_ids = df["source_id"].tolist()
    target_ids = df["target_id"].tolist()
    source_names = df["source_name"].tolist()
    target_names = df["target_name"].tolist()

    genes = []
    for i in range(len(source_ids)):
        if source_ids[i].startswith("ENTREZ:"):
            genes.append(source_names[i])
        if target_ids[i].startswith("ENTREZ:"):
            genes.append(target_names[i])
    return list(set(genes))


def check_sheet_exists(file_name, sheet_name):
    """Check if the sheet exists in the excel file.

    Args:
        file_name (str): excel file name
        sheet_name (str): sheet name

    Returns:
        bool: True if the sheet exists, False otherwise
    """
    workbook = load_workbook(file_name, read_only=True)
    return sheet_name in workbook.sheetnames


def convert_drugbank_to_mesh(drugbank_ids):
    # mychem.info API URL
    url = "https://mychem.info/v1/query"

    # Dictionary to hold DrugBank to MeSH ID mappings
    mapping = {}

    for i in range(0, len(drugbank_ids), 100):
        # Prepare the query
        q = ",".join(drugbank_ids[i : i + 100])
        params = {
            "q": q,
            "fields": "ginas.xrefs,pharmgkb.xrefs.mesh,umls.mesh",
            "scopes": "drugbank.id",
        }

        # Send the request
        response = requests.post(url, params=params)

        # Check if the response is valid
        print(response.status_code, response.text)
        results = response.json()
        for result in results:
            drugbank_id = result["query"]
            xrefs = result.get("ginas", {}).get("xrefs", [])
            xrefs = [xref for xref in xrefs if xref.startswith("MESH:")]
            if len(xrefs) > 0:
                mesh_id = xrefs[0]
                mapping[drugbank_id] = mesh_id
            else:
                xrefs = result.get("pharmgkb", {}).get("xrefs", {}).get("mesh", [])
                if len(xrefs) > 0:
                    mesh_id = xrefs[0]
                    mapping[drugbank_id] = mesh_id
                else:
                    xrefs = result.get("umls", {}).get("mesh", [])
                    if len(xrefs) > 0:
                        mesh_id = xrefs[0]
                        mapping[drugbank_id] = mesh_id

    return mapping


def get_mesh_id4drugbank_id(drugs, output_file, entity_file):
    entities = pd.read_csv(entity_file, sep="\t", dtype=str)

    if "drug_mesh_id" not in drugs.columns:
        for i, row in drugs.iterrows():
            drug_name = row["drug_name"]
            drug_id = row["drug_id"]
            mesh_id = None

            # if the drug list already contains MESH ID, use it directly
            if drug_id.startswith("MESH:"):
                mesh_id = drug_id
            else:
                # otherwise, find from the entity list
                xrefs = entities[entities["id"] == drug_id]["xrefs"].tolist()
                # xrefs might be a list of nan
                xrefs = [xref for xref in xrefs if str(xref) != "nan"]
                print(f"Processing {drug_name} ({drug_id})..., xrefs: {xrefs}")
                if len(xrefs) > 0:
                    xrefs = xrefs[0]
                    xrefs = xrefs.split("|")
                    for xref in xrefs:
                        if xref.startswith("MESH:"):
                            mesh_id = xref
                            break

            drugs.loc[i, "drug_mesh_id"] = mesh_id

    filtered_drugs = drugs[drugs["drug_mesh_id"].notnull()]
    rest_drugs = drugs[drugs["drug_mesh_id"].isnull()]

    print(f"Filtered drugs: {len(filtered_drugs)}, rest drugs: {len(rest_drugs)}")
    # if there are some drugs without MESH ID, use mychem.info API to get it
    if len(rest_drugs) > 0:
        drugbank_ids = rest_drugs["drug_id"].tolist()
        drugbank_ids = [drugbank_id.split(":")[1] for drugbank_id in drugbank_ids]
        mapping = convert_drugbank_to_mesh(drugbank_ids)
        for i, row in rest_drugs.iterrows():
            drug_id = row["drug_id"].split(":")[1]
            mesh_id = mapping.get(drug_id, None)
            rest_drugs.loc[i, "drug_mesh_id"] = mesh_id

    drugs = pd.concat([filtered_drugs, rest_drugs], axis=0)
    drugs.rename(
        columns={
            "drug_id": "target_raw_id",
            "drug_name": "target_name",
            "drug_mesh_id": "target_id",
        },
        inplace=True,
    )
    drugs.to_csv(output_file, sep="\t", index=False)


def save_df(df: pd.DataFrame, output_file, table_name=""):
    """Save the dataframe to the output file.

    Args:
        df (DataFrame): dataframe to save
        output_file (str): output file path
        table_name (str, optional): table name. Defaults to "".
    """
    mode = "w"
    if_sheet_exists = None
    if os.path.exists(output_file):
        mode = "a"
        if_sheet_exists = "replace"

    with pd.ExcelWriter(
        output_file, mode=mode, if_sheet_exists=if_sheet_exists
    ) as writer:
        if table_name:
            # Save the dataframe to a file
            df.to_excel(writer, sheet_name=table_name, index=False)
        else:
            # Save the dataframe to a file
            df.to_excel(writer, index=False)


def kge_score_fn(head, relation, tail, gamma=12.0, model: str = "TransE_l2"):
    """KGE score function.

    Args:
        head (array): head embedding
        relation (array): relation embedding
        tail (array): tail embedding
        gamma (float): gamma
        model (str, optional): model name. Defaults to TransE_l2.

    Returns:
        float: score
    """
    head = th.tensor(head, dtype=th.float32)
    rel = th.tensor(relation, dtype=th.float32)
    tail = th.tensor(tail, dtype=th.float32)
    logsigmoid = fn.logsigmoid
    exp = th.exp

    if model == "TransE_l1":
        score = head + rel - tail
        return exp(logsigmoid(gamma - th.norm(score, p=1, dim=-1))).item()
    elif model == "TransE_l2":
        score = head + rel - tail
        return exp(logsigmoid(gamma - th.norm(score, p=2, dim=-1))).item()
    elif model == "ComplEx":
        real_head, img_head = th.chunk(head, 2, dim=-1)
        real_tail, img_tail = th.chunk(tail, 2, dim=-1)
        real_rel, img_rel = th.chunk(rel, 2, dim=-1)

        score = (
            real_head * real_rel * real_tail
            + img_head * real_rel * img_tail
            + real_head * img_rel * img_tail
            - img_head * img_rel * real_tail
        )
        return exp(logsigmoid(th.sum(score, dim=-1))).item()
    elif model == "DistMult":
        score = head * rel * tail
        return exp(logsigmoid(th.sum(score, dim=-1))).item()
    else:
        raise ValueError("Unknown model")


def cosine_similarity(vec1, vec2):
    return np.dot(vec1, vec2) / (np.linalg.norm(vec1) * np.linalg.norm(vec2))


def euclidean_distance(vec1, vec2):
    return np.sqrt(np.sum(np.square(vec1 - vec2)))


def manhattan_distance(vec1, vec2):
    return np.sum(np.abs(vec1 - vec2))


def pathway_enrichment(gene_entrez_ids):
    # More details about g:Profiler: https://biit.cs.ut.ee/gprofiler/gost
    gp = GProfiler(return_dataframe=True)

    # perform pathway enrichment analysis
    enrichment_results = gp.profile(
        organism="hsapiens",  # for example, human
        query=gene_entrez_ids,
        sources=["KEGG", "GO", "REAC", "WP"],
        all_results=True,
    )  # 返回所有结果

    # filter and display the important results
    significant_results = enrichment_results[enrichment_results["p_value"] < 0.05]
    return significant_results


def create_graph(kg_df):
    # create a new directed graph
    G = nx.DiGraph()

    # assume you have a DataFrame 'kg_df', which contains your knowledge graph data
    # add nodes and edges to the graph
    for _, row in kg_df.iterrows():
        source = (row["source_id"], row["source_type"])
        target = (row["target_id"], row["target_type"])
        G.add_node(source, type=row["source_type"], name=row["source_name"])
        G.add_node(target, type=row["target_type"], name=row["target_name"])
        G.add_edge(
            source,
            target,
            relation=row["relation_type"],
        )

    return G


def load_graph(graph_file):
    G = gt.load_graph(graph_file)
    return G


def create_graph_tool_graph(kg_df):
    # create a new directed graph
    G = gt.Graph(directed=True)

    # define the attributes. These will store the information of each vertex and edge
    v_id = G.new_vertex_property("string")
    v_type = G.new_vertex_property("string")
    v_name = G.new_vertex_property("string")
    e_relation = G.new_edge_property("string")

    # create a dictionary to reference the created vertices, avoid adding the same vertices repeatedly
    vertex_dict = {}

    # add nodes and edges to the graph
    for _, row in kg_df.iterrows():
        # check if the node exists, if not, create it
        if row["source_id"] not in vertex_dict:
            source_vertex = G.add_vertex()
            v_id[source_vertex] = row["source_id"]
            v_type[source_vertex] = row["source_type"]
            v_name[source_vertex] = row["source_name"]
            vertex_dict[row["source_id"]] = source_vertex
        else:
            source_vertex = vertex_dict[row["source_id"]]

        if row["target_id"] not in vertex_dict:
            target_vertex = G.add_vertex()
            v_id[target_vertex] = row["target_id"]
            v_type[target_vertex] = row["target_type"]
            v_name[target_vertex] = row["target_name"]
            vertex_dict[row["target_id"]] = target_vertex
        else:
            target_vertex = vertex_dict[row["target_id"]]

        # add edge
        edge = G.add_edge(source_vertex, target_vertex)
        e_relation[edge] = row["relation_type"]

    # associate the attributes with the graph
    G.vertex_properties["id"] = v_id
    G.vertex_properties["type"] = v_type
    G.vertex_properties["name"] = v_name
    G.edge_properties["relation"] = e_relation

    return G


def find_connected_symptoms_gt(
    drug_ids: List[str], symptoms_set: list, graph: gt.Graph, allowed_disease_ids: list
) -> pd.DataFrame:
    drug_disease_symptom_triples = []
    id_to_vertex = {f"{graph.vp.type[v]}:{graph.vp.id[v]}": v for v in graph.vertices()}

    for drug_id in drug_ids:
        # Find the vertex for the drug
        drug_vertex = None
        drug_name = ""
        drug_composed_id = f"Compound:{drug_id}"
        if id_to_vertex.get(drug_composed_id):
            drug_vertex = id_to_vertex[drug_composed_id]
            drug_name = graph.vp.name[drug_vertex]
        else:
            continue

        # iterate through all symptoms
        for symptom_id in symptoms_set:
            # Find the vertex for the symptom
            symptom_name = ""
            symptom_vertex = None
            symptom_composed_id = f"Symptom:{symptom_id}"
            if id_to_vertex.get(symptom_composed_id):
                symptom_vertex = id_to_vertex[symptom_composed_id]
                symptom_name = graph.vp.name[symptom_vertex]
            else:
                continue

            # check if there is a path
            distance = gt.shortest_distance(
                graph, source=drug_vertex, target=symptom_vertex
            )

            if distance < 3:
                # get all paths
                paths = gt.all_paths(
                    graph, source=drug_vertex, target=symptom_vertex, cutoff=2
                )

                # check if there is at least one disease node in the path
                for path in paths:
                    print(
                        "Path: ", [f"{graph.vp.name[v]}:{graph.vp.id[v]}" for v in path]
                    )
                    for node in path:
                        disease_id = graph.vp.id[node]
                        disease_name = graph.vp.name[node]
                        if (
                            graph.vp.type[node] == "Disease"
                            and disease_id in allowed_disease_ids
                        ):
                            triple = {
                                "drug_id": drug_id,
                                "drug_name": drug_name,
                                "disease_id": disease_id,
                                "disease_name": disease_name,
                                "symptom_id": symptom_id,
                                "symptom_name": symptom_name,
                            }
                            drug_disease_symptom_triples.append(triple)
                            print("Triple: ", triple)

    return pd.DataFrame(drug_disease_symptom_triples).drop_duplicates()


def find_connected_symptoms(
    drug_id, symptoms_set, graph: nx.DiGraph
) -> Tuple[list, dict]:
    connected_symptoms = []
    diseases = {}

    drug = (drug_id, "Compound")
    # iterate through all symptoms
    for symptom_id in symptoms_set:
        # check if there is a path
        symptom = (symptom_id, "Symptom")
        if nx.has_path(graph, drug, symptom):
            # get all paths
            paths = nx.all_simple_paths(graph, drug, symptom, cutoff=3)
            # check if there is at least one disease node in the path
            for path in paths:
                for node in path:
                    n = graph.nodes[node]
                    if n["type"] == "Disease":
                        if symptom_id not in diseases:
                            diseases[symptom_id] = []
                        diseases[symptom_id].append({"id": node[0], "name": n["name"]})
                        connected_symptoms.append(symptom_id)

    unique_connected_symptoms = set(connected_symptoms)

    return unique_connected_symptoms, diseases


def format_name(df, column_name):
    df[column_name] = df[column_name].str.replace("_", " ")
    df[column_name] = df[column_name].str.title()
    return df


def compute_top_n_similar_entities(
    entity_embedding,
    other_entity_embeddings,
    relation_embedding,
    top_n=100,
    model="TransE_l2",
    gamma=12.0,
) -> Tuple[List[int], List[float]]:
    scores = np.array(
        [
            kge_score_fn(
                entity_embedding, relation_embedding, tail, model=model, gamma=gamma
            )
            for tail in other_entity_embeddings
        ]
    )

    logger.info("Get the top n similar entities")
    top_n_indices = np.argsort(scores)[::-1][:top_n]
    top_n_scores = scores[top_n_indices]
    return top_n_indices, top_n_scores


def kge_score_fn_batch(
    heads: List[np.ndarray],
    relation: np.ndarray,
    tails: List[np.ndarray],
    gamma=12.0,
    model="TransE_l2",
):
    """KGE score function for batch processing.

    Args:
        heads (array): batch of head embeddings
        relation (array): relation embedding
        tails (array): batch of tail embeddings
        gamma (float): gamma
        model (str, optional): model name. Defaults to TransE_l2.

    Returns:
        array: batch of scores
    """
    logger.info(f"The type of heads: {type(heads)}, len(heads): {len(heads)}")
    logger.info(f"The type of tails: {type(tails)}, len(tails): {len(tails)}")
    logger.info(f"The type of relation: {type(relation)}, len(relation): {len(relation)}")

    if not isinstance(tails, list) or not all(
        isinstance(tail, np.ndarray) for tail in tails
    ):
        raise ValueError("Tails must be a list of numpy arrays")

    if not isinstance(heads, list) or not all(
        isinstance(head, np.ndarray) for head in heads
    ):
        raise ValueError("Heads must be a list of numpy arrays")

    if len(tails) == 0 or len(heads) == 0:
        return np.array([])

    tail_tensor = th.tensor(np.stack(tails), dtype=th.float32)
    head_tensor = th.tensor(np.stack(heads), dtype=th.float32)
    rel_tensor = th.tensor(relation, dtype=th.float32)

    logsigmoid = fn.logsigmoid
    exp = th.exp

    if model == "TransE_l1":
        score = head_tensor + rel_tensor - tail_tensor
        return exp(logsigmoid(gamma - th.norm(score, p=1, dim=-1))).detach().numpy()
    elif model == "TransE_l2":
        score = head_tensor + rel_tensor - tail_tensor
        return exp(logsigmoid(gamma - th.norm(score, p=2, dim=-1))).detach().numpy()
    elif model == "ComplEx":
        real_head, img_head = th.chunk(head_tensor, 2, dim=-1)
        real_tail, img_tail = th.chunk(tail_tensor, 2, dim=-1)
        real_rel, img_rel = th.chunk(rel_tensor, 2, dim=-1)

        score = (real_head * real_rel - img_head * img_rel) * real_tail + (
            real_head * img_rel + img_head * real_rel
        ) * img_tail
        return exp(logsigmoid(th.sum(score, dim=-1))).detach().numpy()
    elif model == "DistMult":
        score = head_tensor * rel_tensor * tail_tensor
        return exp(logsigmoid(th.sum(score, dim=-1))).detach().numpy()
    else:
        raise ValueError("Unknown model")


def compute_top_n_similar_entities_vectorized(
    entity_embedding: np.ndarray,
    other_entity_embeddings: List[np.ndarray],
    relation_embedding: np.ndarray,
    top_n=100,
    model="TransE_l2",
    gamma=12.0,
) -> Tuple[List[int], List[float]]:
    # calculate all scores
    scores = kge_score_fn_batch(
        [entity_embedding],
        relation_embedding,
        other_entity_embeddings,
        model=model,
        gamma=gamma,
    )

    # get the top n similar entities
    logger.info("Get the top n similar entities")
    top_n_indices = np.argsort(scores)[::-1][:top_n]
    top_n_scores = scores[top_n_indices]
    return top_n_indices, top_n_scores


def convert_id_to_umls(id, id_type, api_key):
    """
    Convert a ID to UMLS ID using BioPortal's REST API.

    :param id: The ID to convert.
    :param id_type: The type of ID to convert. Must be one of MESH, SNOMEDCT, SYMP, MEDDRA.
    :param api_key: Your BioPortal API key.
    :return: The corresponding UMLS ID, if found.
    """
    base_url = "http://data.bioontology.org"
    headers = {"Authorization": f"apikey token={api_key}"}

    # More details on the API here: https://data.bioontology.org/documentation#Class
    # You can get the related UMLS ids for SYMP from the downloaded file here: https://bioportal.bioontology.org/ontologies/SYMP?p=summary
    if id_type not in ["MESH", "SNOMEDCT", "MEDDRA"]:
        print(
            f"Error: {id_type} is not a valid ID type, must be one of MESH, SNOMEDCT, MEDDRA"
        )
        return None

    if id_type in ["MESH", "SNOMEDCT", "MEDDRA"]:
        path = f"http%3A%2F%2Fpurl.bioontology.org%2Fontology%2F{id_type}%2F{id}"

    url = f"{base_url}/ontologies/{id_type}/classes/{path}"
    print("The URL is: ", url)

    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        data = response.json()
        print(json.dumps(data, indent=2))
        mappings = data.get("cui", [])
        if len(mappings) > 0:
            return mappings[0]
        else:
            print(f"Error: No mappings found for {id}")
            return None
    else:
        print(f"Error: {response.status_code}")
        return None
